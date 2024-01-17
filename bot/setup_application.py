import os
import re
import zipfile

from bot.handlers.error_handler import send_stacktrace_to_tg_chat

from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, \
    ReplyKeyboardRemove
from telegram.ext import ContextTypes, ApplicationBuilder, CommandHandler, MessageHandler, ConversationHandler, \
    CallbackQueryHandler, Application
from telegram.ext.filters import TEXT, Document, COMMAND

from bot.parser.parse_pdf import find_info, check_pdf_to_be_valid_doc, get_floors_pics, floors, process_pdf, \
    write_to_table, cad_id_from_table
from bot.parser.tables.temp_writer import write_to_table, write_first_and_seventh, write_first_and_seventh_to_dir
from bot.parser.tables.parse_xlsx import get_addr_and_cad_id, make_register_file, check_for_spaces_in_column, \
    join_owners_and_registry

from bot.parser.parse_website.util import Parser

from bot.states import MainDialogStates, ApiDialogStates

import logging
from logging.handlers import RotatingFileHandler

from bot.middlewares import SessionMiddleware, UserMiddleware, Middleware

from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from database import engine

from crud import request as request_service

from client_server_api import client_server_api

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        RotatingFileHandler("logs/bot.log", maxBytes=20000, backupCount=5),
        logging.StreamHandler(),
    ]
)
logging.getLogger("httpx").setLevel(logging.WARNING)

logger = logging.getLogger(__name__)

token = os.environ.get("TOKEN")
order_id = 0

main_keyboard = ReplyKeyboardMarkup([
    [KeyboardButton("ЕГРН"), KeyboardButton("Импорт ФИО")]],
    input_field_placeholder='Выберите действие:',
    resize_keyboard=True,
    one_time_keyboard=True,
)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = await context.bot.send_message(chat_id=update.effective_chat.id,
                                             text="Выберите действие из меню",
                                             reply_markup=main_keyboard)
    context.user_data['messages_to_delete'] = []
    context.user_data["messages_to_delete"].extend([message, update.message])


async def egrn_chose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "ЕГРН":
        remove_markup = ReplyKeyboardRemove()
        message = await context.bot.send_message(update.message.chat.id, "Вставьте файл выписки на здание:",
                                                 reply_markup=remove_markup)
        context.user_data["messages_to_delete"] = []
        context.user_data["messages_to_delete"].extend([message, update.message])
        return MainDialogStates.GET_DOC
    else:
        print("transfer")
        return await test_fio(update, context)


async def get_doc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = await context.bot.get_file(update.message.document)

    path_prefix_for_received_files = 'bot/files/received/'

    message_to_reply = update.message

    document_name = update.message.document.file_name
    src = path_prefix_for_received_files + document_name
    await file.download_to_drive(src)

    if not document_name.endswith(".pdf"):  # file format check
        if not document_name.endswith("zip"):
            message = await update.message.reply_text("Неверный формат файла, попробуйте ещё раз")
            context.user_data["messages_to_delete"].extend([message, update.message])
            return

        with open(src, 'rb') as zip:
            if zipfile.is_zipfile(zip):
                z = zipfile.ZipFile(zip)
                for member in z.namelist():
                    if member.split('.')[-1] == 'pdf':
                        z.extract(member=member, path=path_prefix_for_received_files)
                        new_name = find_info(path_prefix_for_received_files + member)
                        new_name = ' '.join(new_name.values()) + ".pdf"
                        src = path_prefix_for_received_files + new_name
                        os.rename(path_prefix_for_received_files + member, src)
                        message_to_reply = await context.bot.send_document(update.message.chat.id, src)
                        await update.message.delete()
                        break

    if not check_pdf_to_be_valid_doc(src):  # file contains specific title
        message = await update.message.reply_text("Неверный формат файла, попробуйте ещё раз")
        context.user_data["messages_to_delete"].extend([message, update.message])
        os.remove(src)
        return

    context.user_data["src"] = src

    tmp_msg = await update.effective_message.reply_text(
        "Обработка начата, ожидайте..."
    )
    context.user_data["messages_to_delete"].extend([tmp_msg])

    doc = write_first_and_seventh(*[*process_pdf(src), src])
    context.user_data["fs_doc"] = doc

    cad = find_info(src)["Кадастровый номер"]
    context.user_data["cad"] = cad

    cad_id, addr = get_addr_and_cad_id(doc)

    filename = f"ЕГРН {cad_id} {addr}".replace('/', '_').replace(':', '') + ".xlsx"

    reply = f'''МКД: {addr}\nКадастровый номер: {cad_id}\n\nСоздать файл РеестрМКД с данными из онлайн-справки Росреестра?'''

    pics = InlineKeyboardButton(text="Планы", callback_data="plans")
    makefile = InlineKeyboardButton(text="РеестрМКД", callback_data="makefile")
    markup = InlineKeyboardMarkup([[pics, makefile]])

    await message_to_reply.reply_document(doc, quote=True, filename=filename)

    await update.effective_message.reply_text(text=reply, reply_markup=markup)

    await delete_messages(context)
    await delete_message_or_skip(update.effective_message)
    return MainDialogStates.CHOOSE_OPTION


async def online_chose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    text = update.callback_query.data
    if text == "makefile":
        message = await context.bot.send_message(update.callback_query.message.chat.id,
                                                 "Обрабатываю файл, ожидайте...")
        context.user_data["messages_to_delete"].append(message)

        src = context.user_data["fs_doc"]
        try:
            parser_obj = Parser(src)
            await parser_obj.process_objects()
        except Exception as e:
            message = await update.effective_message.reply_text(text="Ошибка при обработке файла")
            logger.error(e)
            context.user_data["messages_to_delete"].append(message)
            parser_obj = Parser(src)
            await parser_obj.process_objects()

        doc, area = make_register_file(src)

        context.user_data["res_file"] = doc

        reply = f'Площадь КВ+НЖ+ММ = {area:.1f} кв.м.\nЗагрузить ФИО собственников в РеестрМКД?'

        table_button = InlineKeyboardButton(text="Таблица", callback_data="table")
        extract_button = InlineKeyboardButton(text="Выписка", callback_data="extract")
        not_button = InlineKeyboardButton(text="Нет", callback_data="no")

        markup = InlineKeyboardMarkup([[table_button, extract_button, not_button]])

        message = await update.effective_message.reply_text(text=reply, reply_markup=markup)

        context.user_data["messages_to_delete"].append(message)

        return MainDialogStates.CHOOSE_OPTION_REGISTRY


async def choose_option_registry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    text = update.callback_query.data
    if text == "table":
        message = await context.bot.send_message(update.callback_query.message.chat.id,
                                                 "Вставьте файл с ФИО в формате .xlsx")
        context.user_data["messages_to_delete"].extend([update.message, message])
        return MainDialogStates.ASKED_TABLE
    if text == "no":
        doc = context.user_data["res_file"]

        await delete_messages(context)
        await update.effective_message.reply_document(doc, reply_markup=main_keyboard)
        return ConversationHandler.END
    if text == "extract":
        message = await update.effective_message.reply_text("Функционал в разработке")
        context.user_data["messages_to_delete"].extend([update.message, message])
        return MainDialogStates.CHOOSE_OPTION_REGISTRY


letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
letters_keyboard = ['A(1)', 'B(2)', 'C(3)', 'D(4)', 'E(5)', 'F(6)', 'G(7)', 'H(8)', 'I(9)', 'J(10)']


def make_letters_keyboard(mapping: dict[str, str], skip_button: bool = False) -> InlineKeyboardMarkup:
    columns = mapping
    buttons_in_row = 5
    letters_to_iterate = [letters_keyboard[i:i + buttons_in_row]
                          for i in range(0, len(letters_keyboard), buttons_in_row)]
    buttons = [[InlineKeyboardButton(text=x, callback_data=x)
                for x in y
                if x[0] not in columns.values()]
               for y in letters_to_iterate]
    if skip_button:
        buttons.append([InlineKeyboardButton(text="Пропустить", callback_data="0")])
    return InlineKeyboardMarkup(buttons)


async def got_table(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print("i'm here")

    file = await context.bot.get_file(update.message.document)
    received_files_path = 'bot/files/received/'

    document_name = update.message.document.file_name
    src = received_files_path + document_name
    await file.download_to_drive(src)

    context.user_data["owners_file"] = src
    columns = {}

    markup = make_letters_keyboard(columns)

    message = await update.effective_message.reply_text("Укажите номер колонки с кадастровым номером",
                                                        reply_markup=markup)
    context.user_data["messages_to_delete"].append(message)
    context.user_data["previous_column"] = [message]
    context.user_data["columns"] = columns
    return MainDialogStates.ASKED_CAD


async def test_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print("got")
    if update.message.text == 'Импорт ФИО':
        print("done")
        await update.message.reply_text("Вставьте пустой РеестрМКД в формате .xlsx")
        context.user_data["messages_to_delete"] = []
        return MainDialogStates.FIO_TEST_ASKED_REGISTRY


async def test_fio_got_registry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print("test fio got registry")
    file = await context.bot.get_file(update.message.document)

    received_files_path = 'bot/files/received/'

    document_name = update.message.document.file_name
    src = received_files_path + document_name
    await file.download_to_drive(src)

    context.user_data["res_file"] = src
    await update.effective_message.reply_text("Вставьте файл с ФИО в формате .xlsx")
    return MainDialogStates.ASKED_TABLE


async def asked_cad(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['cad_id'] = text[0]
        await delete_previous_column(context)

        message = await update.effective_message.reply_text(f"Кадастровый номер: {text}")
        context.user_data["previous_column"].append(message)

        markup = make_letters_keyboard(columns)

        message = await update.effective_message.reply_text("Укажите номер колонки с ФИО собственников",
                                                            reply_markup=markup)
        context.user_data["previous_column"].append(message)
        context.user_data["columns"] = columns
        return MainDialogStates.ASKED_NAMES


async def delete_previous_column(context: ContextTypes.DEFAULT_TYPE):
    for message in context.user_data.get("previous_column", []):
        await message.delete()
    context.user_data["previous_column"] = []


async def asked_owners(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # (если в ячейке нет пробелов, то "укажите дополнительные колонки для Имени и Отчества (через запятую)")
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['owner'] = text[0]

        await delete_previous_column(context)

        message = await update.effective_message.reply_text(f"ФИО собственников: {text}")
        context.user_data["previous_column"].append(message)

        markup = make_letters_keyboard(columns)

        # send the same with update.effective_message.reply_text
        message = await update.effective_message.reply_text(
            "Укажите номер колонки с Видом, № и датой госрегистрации:",
            reply_markup=markup)

        context.user_data["previous_column"].append(message)
        context.user_data["columns"] = columns
        return MainDialogStates.ASKED_REG


async def asked_registration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['registred'] = text[0]
        await delete_previous_column(context)
        message = await update.effective_message.reply_text(f"Вид, № и датой госрегистрации: {text}")
        context.user_data["previous_column"].append(message)

        markup = make_letters_keyboard(columns, skip_button=True)

        # send the same with update.effective_message.reply_text
        message = await update.effective_message.reply_text(
            "Укажите номер колонки с № и датой выписки ЕГРН",
            reply_markup=markup)

        context.user_data["previous_column"].append(message)
        context.user_data["columns"] = columns
        return MainDialogStates.ASKED_EXTRACT


async def asked_extract(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Введите или укажите номер колонки с указанием доли в помещении:
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['extract'] = text[0]
        await delete_previous_column(context)

        message = await update.effective_message.reply_text(f"№ и дата выписки ЕГРН: {text}")
        context.user_data["previous_column"].append(message)

        markup = make_letters_keyboard(columns)

        message = await context.bot.send_message(update.callback_query.message.chat.id,
                                                 "Укажите номер колонки с указанием доли в помещении:",
                                                 reply_markup=markup)

        context.user_data["previous_column"].append(message)
        context.user_data["columns"] = columns
        return MainDialogStates.ASKED_PARTS
    elif text == "0":
        columns = context.user_data["columns"]

        markup = make_letters_keyboard(columns)

        await delete_previous_column(context)

        message = await context.bot.send_message(update.callback_query.message.chat.id,
                                                 "Укажите номер колонки с указанием доли в помещении:",
                                                 reply_markup=markup)
        context.user_data["previous_column"].append(message)
        return MainDialogStates.ASKED_PARTS


async def asked_parts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['part'] = text[0]
        await delete_previous_column(context)

        regisrtry_file = context.user_data["res_file"]
        owners_file = context.user_data["owners_file"]

        regisrtry_file = join_owners_and_registry(regisrtry_file, owners_file, columns)

        await context.bot.send_document(update.callback_query.message.chat.id, regisrtry_file,
                                        reply_markup=main_keyboard)
        await delete_messages(context)
        return MainDialogStates.CHOOSE_OPTION


async def pics_chose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.callback_query.data
    if text == "plans":
        src = context.user_data["src"]
        cad = context.user_data["cad"]
        floors_reply = floors(src)

        context.user_data["available_pages"] = floors_reply[1]
        floors_reply = floors_reply[0]

        floors_file = f'bot/files/{cad}.txt'
        with open(floors_file, 'wt', encoding='utf-8') as f:
            f.write(floors_reply)

        context.user_data["floors_file"] = floors_file

        message = await context.bot.send_message(update.effective_message.chat.id, floors_reply)
        context.user_data["messages_to_delete"].append(message)

        message = await context.bot.send_message(update.callback_query.message.chat.id,
                                                 "Введите номера листов для получения архива поэтажных планов: (через запятую или пробел)")
        context.user_data["messages_to_delete"].append(message)
        return MainDialogStates.FLOORS_PICS


async def floor_pics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    floors_numbers = re.split(', |,| ', update.message.text)
    available_pages = context.user_data["available_pages"]
    pages_to_check = []
    for y in available_pages:
        pages_to_check.extend(y)
    for floor in floors_numbers:
        if int(floor) not in pages_to_check:
            print(type(floor), pages_to_check)
            message = await update.effective_message.reply_text(
                "Вы ввели некорректные номера листов. Пожалуйста, введите номера из списка выше")
            context.user_data["messages_to_delete"].append(message)
            return MainDialogStates.FLOORS_PICS

    src = context.user_data["src"]
    floors_filename = context.user_data.get("floors_file", None)
    res = get_floors_pics(src, floors_numbers, floors_filename)
    if res:
        await context.bot.send_document(update.message.chat.id, res, reply_markup=main_keyboard)
        context.user_data["messages_to_delete"].append(update.message)
        await delete_messages(context)
    return MainDialogStates.CHOOSE_OPTION


async def delete_message_or_skip(message) -> bool:
    try:
        await message.delete()
        return True
    except Exception as e:
        return False


async def delete_messages(context: ContextTypes.DEFAULT_TYPE):
    for message in context.user_data.get("messages_to_delete", []):
        await delete_message_or_skip(message)
    context.user_data["messages_to_delete"] = []


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await delete_messages(context)
    await delete_message_or_skip(update.message)
    return ConversationHandler.END


async def api_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    request_id = int(update.callback_query.data.split('_')[-1])
    context.user_data["request_id"] = request_id
    message = await update.effective_message.reply_text("Вставьте .pdf файл выписки")
    context.user_data["messages_to_delete"] = [update.effective_message, message]
    return ApiDialogStates.ASKED_EXTRACT


# noinspection DuplicatedCode
async def api_asked_extract(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = await context.bot.get_file(update.message.document)

    received_files_path = 'bot/files/received/'

    message_to_reply = update.message

    document_name = update.message.document.file_name
    src = received_files_path + document_name
    await file.download_to_drive(src)

    if not document_name.endswith(".pdf"):  # file format check
        if not document_name.endswith("zip"):
            message = await update.message.reply_text("Неверный формат файла, попробуйте ещё раз")
            context.user_data["messages_to_delete"].extend([message, update.message])
            return

        with open(src, 'rb') as zip:
            if zipfile.is_zipfile(zip):
                z = zipfile.ZipFile(zip)
                for member in z.namelist():
                    if member.split('.')[-1] == 'pdf':
                        z.extract(member=member, path=received_files_path)
                        new_name = find_info(received_files_path + member)
                        new_name = ' '.join(new_name.values()) + ".pdf"
                        src = received_files_path + new_name
                        os.rename(received_files_path + member, src)
                        message_to_reply = await context.bot.send_document(update.message.chat.id, src)
                        await update.message.delete()
                        break

    if not check_pdf_to_be_valid_doc(src):  # file contains some title
        message = await update.message.reply_text("Неверный формат файла, попробуйте ещё раз")
        context.user_data["messages_to_delete"].extend([message, update.message])
        os.remove(src)
        return

    request_id = context.user_data["request_id"]
    request = await request_service.get_request(session=context.session, request_id=request_id)

    cad = find_info(src)["Кадастровый номер"]
    context.user_data["cad"] = cad

    cadnum = cad_id_from_table(src)

    if cadnum != request.cadnum:
        message = await update.message.reply_text("Неверный кадастровый номер, попробуйте ещё раз")
        context.user_data["messages_to_delete"].extend([message, update.message])
        os.remove(src)
        return

    context.user_data["src"] = src

    tmp_msg = await update.effective_message.reply_text(
        "Обработка начата, ожидайте..."
    )
    context.user_data["messages_to_delete"].extend([tmp_msg])

    doc = write_first_and_seventh_to_dir(*[*process_pdf(src), f"files"])
    context.user_data["fs_doc"] = doc

    request.r1r7_filename = doc
    # print(doc)
    await context.session.commit()

    cad_id, addr = get_addr_and_cad_id(doc)
    reply = (f'МКД: {addr}\nКадастровый номер: {cad_id}\n\n'
             'Проверьте правильность данных. Если данные неверны, нажмите "Редактировать", если верны - "Подтвердить"')

    edit_button = InlineKeyboardButton(text="Редактировать", callback_data=f"r1r7_edit_{request_id}")
    confirm_button = InlineKeyboardButton(text="Подтвердить", callback_data=f"r1r7_confirm_{request_id}")
    markup = InlineKeyboardMarkup([[edit_button, confirm_button]])

    filename = f"ЕГРН {cad_id} {addr}".replace('/', '_').replace(':', '') + ".xlsx"

    document_message = await message_to_reply.reply_document(doc, quote=True, filename=filename)

    text_message = await update.effective_message.reply_text(text=reply, reply_markup=markup)

    await delete_messages(context)
    await delete_message_or_skip(update.message)

    context.user_data["messages_to_delete"].extend([document_message, text_message])

    return ApiDialogStates.CONFIRM_R1R7


async def api_confirm_r1r7(update: Update, context: ContextTypes.DEFAULT_TYPE):
    request_id = context.user_data["request_id"]
    request = await request_service.get_request(session=context.session, request_id=request_id)

    r1r7_filename = request.r1r7_filename
    room_rows_count = get_num_of_rows_in_xlsx(r1r7_filename, 'Раздел 7')

    fio_rows_count = 0

    if request.fio_is_provided:
        fio_file = client_server_api.get_fio_file(request.order_id)
        fio_rows_count = get_num_of_rows_in_xlsx(fio_file)

    await client_server_api.post_request(
        request.order_id,
        'r1r7_is_ready',
        room_rows_count=room_rows_count,
        fio_rows_count=fio_rows_count
    )

    await delete_messages(context)

    await update.callback_query.answer(
        "Данные подтверждены. Отправлено в клиентского бота",
        show_alert=True,
    )

    await delete_messages(context)

    return ConversationHandler.END


async def api_ask_edit_r1r7(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.effective_message.reply_text("Вставьте исправленный файл Р1Р7")
    return ApiDialogStates.EDIT_R1R7


async def api_edit_r1r7(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = await context.bot.get_file(update.message.document)

    received_files_path = 'bot/files/received/'

    document_name = update.message.document.file_name

    if not document_name.endswith(".xlsx"):
        message = await update.message.reply_text("Неверный формат файла, попробуйте ещё раз")
        context.user_data["messages_to_delete"].extend([message, update.message])
        return

    src = received_files_path + document_name
    await file.download_to_drive(src)

    await api_confirm_r1r7(update, context)

    return ConversationHandler.END


async def api_confirm_registry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        await update.callback_query.answer(
            text="Результат отправлен",
            show_alert=True
        )
        request_id = int(update.callback_query.data.split("_")[-1])
    else:
        await update.effective_message.reply_text("Данные подтверждены")
        request_id = context.user_data["request_id"]

    request = await request_service.get_request(session=context.session, request_id=request_id)

    await client_server_api.post_request(request.order_id, 'registry_is_ready', total_area=123.45)

    await delete_messages(context)
    return ConversationHandler.END


async def api_ask_edit_registry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    request_id = int(update.callback_query.data.split("_")[-1])
    context.user_data["request_id"] = request_id

    await update.effective_message.reply_text("Вставьте исправленный файл реестра")

    return ApiDialogStates.EDIT_REGISTRY


async def api_edit_registry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = await context.bot.get_file(update.message.document)

    # received_files_path = 'files/'

    document_name = update.message.document.file_name

    if not document_name.endswith(".xlsx"):
        message = await update.message.reply_text("Неверный формат файла, попробуйте ещё раз")
        context.user_data["messages_to_delete"].extend([message, update.message])
        return

    request = await request_service.get_request(session=context.session, request_id=context.user_data["request_id"])

    src = request.registry_filename
    await file.download_to_drive(src)

    await api_confirm_registry(update, context)


async def api_ask_insert_owners(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    request_id = int(update.callback_query.data.split("_")[-1])
    context.user_data["request_id"] = request_id

    request = await request_service.get_request(session=context.session, request_id=context.user_data["request_id"])

    if request.fio_is_provided:
        columns = {}

        markup = make_letters_keyboard(columns)

        message = await update.effective_message.reply_text("Укажите номер колонки с кадастровым номером",
                                                            reply_markup=markup)
        context.user_data["messages_to_delete"].append(message)
        context.user_data["previous_column"] = [message]
        context.user_data["columns"] = columns
        return ApiDialogStates.ASKED_CAD_COLUMN


async def api_asked_cad_column(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['cad_id'] = text[0]
        await delete_previous_column(context)

        message = await update.effective_message.reply_text(f"Кадастровый номер: {text}")
        context.user_data["previous_column"].append(message)

        markup = make_letters_keyboard(columns)

        message = await update.effective_message.reply_text("Укажите номер колонки с ФИО собственников",
                                                            reply_markup=markup)
        context.user_data["previous_column"].append(message)
        context.user_data["columns"] = columns
        return ApiDialogStates.ASKED_NAMES_COLUMN


async def api_asked_names_column(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # (если в ячейке нет пробелов, то "укажите дополнительные колонки для Имени и Отчества (через запятую)")
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['owner'] = text[0]

        await delete_previous_column(context)

        message = await update.effective_message.reply_text(f"ФИО собственников: {text}")
        context.user_data["previous_column"].append(message)

        markup = make_letters_keyboard(columns)

        # send the same with update.effective_message.reply_text
        message = await update.effective_message.reply_text(
            "Укажите номер колонки с Видом, № и датой госрегистрации:",
            reply_markup=markup)

        context.user_data["previous_column"].append(message)
        context.user_data["columns"] = columns
        return ApiDialogStates.ASKED_REG_COLUMN


async def api_asked_reg_column(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['registred'] = text[0]
        await delete_previous_column(context)
        message = await update.effective_message.reply_text(f"Вид, № и датой госрегистрации: {text}")
        context.user_data["previous_column"].append(message)

        markup = make_letters_keyboard(columns, skip_button=True)

        # send the same with update.effective_message.reply_text
        message = await update.effective_message.reply_text(
            "Укажите номер колонки с № и датой выписки ЕГРН",
            reply_markup=markup)

        context.user_data["previous_column"].append(message)
        context.user_data["columns"] = columns
        return ApiDialogStates.ASKED_EXTRACT_COLUMN


async def api_asked_extract_column(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Введите или укажите номер колонки с указанием доли в помещении:
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['extract'] = text[0]
        await delete_previous_column(context)

        message = await update.effective_message.reply_text(f"№ и дата выписки ЕГРН: {text}")
        context.user_data["previous_column"].append(message)

        markup = make_letters_keyboard(columns)

        message = await context.bot.send_message(update.callback_query.message.chat.id,
                                                 "Укажите номер колонки с указанием доли в помещении:",
                                                 reply_markup=markup)

        context.user_data["previous_column"].append(message)
        context.user_data["columns"] = columns
        return ApiDialogStates.ASKED_PARTS_COLUMN
    elif text == "0":
        columns = context.user_data["columns"]

        markup = make_letters_keyboard(columns)

        await delete_previous_column(context)

        message = await context.bot.send_message(update.callback_query.message.chat.id,
                                                 "Укажите номер колонки с указанием доли в помещении:",
                                                 reply_markup=markup)
        context.user_data["previous_column"].append(message)
        return ApiDialogStates.ASKED_PARTS_COLUMN


async def api_asked_parts_column(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()

    text = update.callback_query.data
    if text in letters_keyboard:
        columns = context.user_data["columns"]

        columns['part'] = text[0]
        await delete_previous_column(context)

        request = await request_service.get_request(context.user_data["request_id"])

        registry_file = request.registry_filename

        owners_file = client_server_api.get_request(request.id)
        owners_filename = f'files/{request.cadnum}.xlsx'
        with open(owners_filename, 'wb') as file:
            file.write(owners_file)

        registry_file = join_owners_and_registry(registry_file, owners_file, columns)

        await context.bot.send_document(update.callback_query.message.chat.id, registry_file,
                                        reply_markup=main_keyboard)
        await delete_messages(context)
        await api_confirm_registry(update, context)


def get_num_of_rows_in_xlsx(filename: str, sheet_name: str = None, count_first: bool = False) -> int:
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    wb: Workbook = load_workbook(filename, read_only=True, data_only=True)
    ws: Worksheet = wb.active if sheet_name is None else wb[sheet_name]
    return ws.max_row if count_first else ws.max_row - 1


def get_application():
    # persistence_input = PersistenceInput(bot_data=True, user_data=True, chat_data=True)
    # persistence = PicklePersistence('bot/bot_data.pickle', store_data=persistence_input, update_interval=1)
    # app = ApplicationBuilder().token(token).persistence(persistence).build()

    session_maker = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
    middleware = Middleware(
        [
            SessionMiddleware(session_maker),
            UserMiddleware(),
        ],
    )

    app = ApplicationBuilder().token(token).build()

    app.add_handler(CommandHandler('start', start))

    conversation_handler = ConversationHandler(
        entry_points=[
            MessageHandler(filters=TEXT & ~COMMAND, callback=egrn_chose),
            # MessageHandler(filters=TEXT & ~COMMAND, callback=test_fio)
        ],
        states={
            MainDialogStates.GET_DOC: [MessageHandler(filters=Document.ALL, callback=get_doc)],
            MainDialogStates.CHOOSE_OPTION: [
                CallbackQueryHandler(callback=pics_chose, pattern="plans"),
                CallbackQueryHandler(callback=online_chose, pattern="makefile"),
            ],
            MainDialogStates.FLOORS_PICS: [MessageHandler(filters=TEXT & ~COMMAND, callback=floor_pics)],
            MainDialogStates.CHOOSE_OPTION_REGISTRY: [
                CallbackQueryHandler(callback=choose_option_registry, pattern="extract"),
                CallbackQueryHandler(callback=choose_option_registry, pattern="no"),
                CallbackQueryHandler(callback=choose_option_registry, pattern="table"),
            ],
            MainDialogStates.FIO_TEST_ASKED_REGISTRY: [
                MessageHandler(filters=Document.ALL, callback=test_fio_got_registry)],
            MainDialogStates.ASKED_TABLE: [MessageHandler(filters=Document.ALL, callback=got_table)],
            MainDialogStates.ASKED_CAD: [CallbackQueryHandler(callback=asked_cad, pattern='^')],
            MainDialogStates.ASKED_NAMES: [CallbackQueryHandler(callback=asked_owners, pattern='^')],
            MainDialogStates.ASKED_REG: [CallbackQueryHandler(callback=asked_registration, pattern='^')],
            MainDialogStates.ASKED_EXTRACT: [CallbackQueryHandler(callback=asked_extract, pattern='^')],
            MainDialogStates.ASKED_PARTS: [CallbackQueryHandler(callback=asked_parts, pattern='^')],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        # persistent=True,
        # name="main_dialog",
    )

    app.add_handler(conversation_handler)

    api_conversation_handler_r1r7 = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(pattern="r1r7_start_", callback=api_start),
        ],
        states={
            ApiDialogStates.ASKED_EXTRACT: [MessageHandler(filters=Document.ALL, callback=api_asked_extract)],
            ApiDialogStates.CONFIRM_R1R7: [
                CallbackQueryHandler(pattern="r1r7_confirm_", callback=api_confirm_r1r7),
                CallbackQueryHandler(pattern="r1r7_edit_", callback=api_ask_edit_r1r7)
            ],
            ApiDialogStates.EDIT_R1R7: [MessageHandler(filters=Document.ALL, callback=api_edit_r1r7)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
        # persistent=True,
        # name="main_dialog",
    )

    app.add_handler(api_conversation_handler_r1r7)

    api_conversation_handler_registry = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(pattern="registry_confirm_", callback=api_confirm_registry),
            CallbackQueryHandler(pattern="registry_edit_", callback=api_ask_edit_registry),
            CallbackQueryHandler(pattern="registry_insert_owners_", callback=api_ask_insert_owners),
        ],
        states={
            ApiDialogStates.EDIT_REGISTRY: [
                MessageHandler(filters=Document.ALL, callback=api_edit_registry)],
            ApiDialogStates.ASKED_CAD_COLUMN: [MessageHandler(filters=TEXT & ~COMMAND, callback=api_asked_cad_column)],
            ApiDialogStates.ASKED_NAMES_COLUMN: [
                MessageHandler(filters=TEXT & ~COMMAND, callback=api_asked_names_column)],
            ApiDialogStates.ASKED_REG_COLUMN: [MessageHandler(filters=TEXT & ~COMMAND, callback=api_asked_reg_column)],
            ApiDialogStates.ASKED_EXTRACT_COLUMN: [
                MessageHandler(filters=TEXT & ~COMMAND, callback=api_asked_extract_column)],
            ApiDialogStates.ASKED_PARTS_COLUMN: [
                MessageHandler(filters=TEXT & ~COMMAND, callback=api_asked_parts_column)],
            ApiDialogStates.CONFIRM_REGISTRY: [
                CallbackQueryHandler(pattern="registry_confirm_", callback=api_confirm_registry),
                CallbackQueryHandler(pattern="registry_edit_", callback=api_ask_edit_registry),
            ],
        },

        fallbacks=[CommandHandler('cancel', cancel)],
        # persistent=True,
        # name="main_dialog",
    )

    app.add_handler(api_conversation_handler_registry)

    app.add_error_handler(send_stacktrace_to_tg_chat)

    middleware.attach_to_application(app)

    return app

from __future__ import annotations

import re
from typing import Tuple, List, Any

from pdfplumber.page import Page
from pdfplumber import open as pdfopen
from itertools import groupby, count
from zipfile import ZipFile
from collections import OrderedDict, _OrderedDictValuesView
import openpyxl


def check_pdf_to_be_valid_doc(filename):
    with pdfopen(filename) as pdffile:
        pat = 'Выписка из Единого государственного реестра недвижимости об объекте недвижимости'
        return pdffile.pages[0].search(pat)


# def extract_pdf_from_zip(src):
#


def address(pages: list[Page]) -> tuple[str, str]:
    _name = 'Адрес'
    tables = pages[0].extract_tables()
    address = tables[3][1][1]
    return _name, address


def cad_id(pages: list[Page]) -> tuple[str, str]:
    _name = 'Кадастровый номер'
    tables = pages[0].extract_tables()
    cad_id = tables[2][1][1]
    return _name, cad_id.replace(':', '')


def cad_id_from_table(src):
    with pdfopen(src) as pdffile:
        pages = pdffile.pages
        tables = pages[0].extract_tables()
        cad_id = tables[2][1][1]
        return cad_id


def resize_columns(ws) -> None:
    for col in ws.columns:
        set_len = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell.value)) > set_len:
                set_len = len(str(cell.value))
        set_col_width = set_len + 5
        ws.column_dimensions[column].width = set_col_width


def process_pdf(src):
    addr, cad = find_info(src).values()
    with pdfopen(src) as pdffile:
        pages = pdffile.pages
        first_div, seventh_div = [], []
        for page in pages:
            tables = page.extract_tables()
            if tables:
                div_id = tables[1][0][0].split()[-1]
                if div_id == '1':
                    if not first_div:
                        first_div.extend(tables[2])
                    first_div.extend(tables[3])
                if div_id == '7':
                    if not seventh_div:
                        seventh_div.extend(tables[3])
                    else:
                        seventh_div.extend(tables[3])

        for row in first_div:
            if row[0] == 'Получатель выписки:':
                row[1] = ''

        return addr, cad, first_div, seventh_div


def write_to_table(addr, cad, first_div, seventh_div):
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.remove(ws)
    ws1 = wb.create_sheet('Раздел 1')
    ws7 = wb.create_sheet('Раздел 7')
    for row in first_div:
        ws1.append([row[0], row[1]])
    for row in seventh_div:
        ws7.append([x for x in row if x])
    resize_columns(ws1)
    resize_columns(ws7)
    res = f'files/xlsx/{cad}_{addr}.xlsx'
    wb.save(res)
    return res


DEFAULT_CHECKS = [address, cad_id]


def find_info(filename: str) -> dict:
    with pdfopen(filename) as pdffile:
        pages = pdffile.pages
        res = {}
        for check in DEFAULT_CHECKS:
            check_res = check(pages)
            res[check_res[0]] = check_res[1]
        return res


def repr_info(info: dict[str, str]) -> str:
    res = ''
    for key, value in info.items():
        res += f'{key}: {value}\n'
    return res


def get_floor(page: Page) -> str:
    tables = page.extract_tables()
    floor_field = tables[2][1][1].split()

    floor = ''.join(floor_field[-1][::2]) if floor_field[0] == 'ННооммеерр' else floor_field[-1]

    floor = floor.split()[-1]
    return floor if floor != '(этажей):' else 'б/н'


def _floors(pages: list[Page]) -> tuple[str, _OrderedDictValuesView[str, list[Any]]]:
    def _str():
        res = ''
        for key, value in pages_by_floor.items():
            groups = groupby(value, lambda n, c=count(): n - next(c))
            value = [list(g) for k, g in groups]

            # groups pages by continuous sequences
            res += f'Этаж {key}: ' \
                   f'Листы {", ".join([str(l[0]) if len(l) <= 1 else f"{l[0]}-{l[-1]}" for l in value])}\n'
        return res

    _name = 'Справка по разделу 8:\n'
    pages_by_floor = OrderedDict()
    for page in pages:
        if page.search('Раздел 8'):
            start_page_number = page.page_number
            break

    for page in pages[start_page_number - 1:]:
        page_number = page.page_number
        floor = get_floor(page)

        if floor not in pages_by_floor:
            pages_by_floor[floor] = []
        pages_by_floor[floor].append(page_number)

    pages_by_floor = OrderedDict(sorted(pages_by_floor.items(), key=lambda x: x[0]))

    return _name + _str(), list(pages_by_floor.values())

# def get_pages_numbers_with_pics(filename)


def floors(filename: str) -> tuple[str, Any]:
    with pdfopen(filename) as pdffile:
        return _floors(pdffile.pages)


def get_floors_pics(filename: str, pages: list[int], floors_filename: str | None) -> str:
    print(filename, pages)
    with pdfopen(filename) as pdffile:
        cad = cad_id(pdffile.pages)
        src = f'bot/files/zips/{cad[1]} поэтажные планы.zip'
        folder_name = re.split('[/.]', src)[2]
        with ZipFile(src, 'w') as zipObj:
            zipObj.mkdir(folder_name)
            for page_number in pages:
                page = pdffile.pages[int(page_number) - 1]
                image = page.images[0]
                zipObj.writestr(f"{folder_name}/{cad[1]} - этаж {get_floor(page)} - лист {page_number}.jpg",
                                image['stream'].get_data())

            with open(floors_filename, 'rt', encoding='utf-8') as f:
                zipObj.writestr(f"{folder_name}/Справка по разделу 8.txt", f.read())
    return src


if __name__ == '__main__':
    src = '../files/received/783100011313013_Пушкинская_улица,_дом_9_report_2a335683_ff0f_4397.pdf'
    print(repr_info(find_info(src)))
    print(get_floors_pics(src, [8, 9, 10]), 'test.zip')

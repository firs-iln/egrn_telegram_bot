import pprint

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

from selenium.common.exceptions import TimeoutException

from concurrent.futures import ThreadPoolExecutor, as_completed

from .parse_web import parse_land, parse_building, parse_room

from egrn_api import api, SearchResponse

import asyncio

"""
Общая информация
    Вид объекта недвижимости Здание
    Статус объекта Актуально
    Кадастровый номер 78:06:0220603:2746
    Дата присвоения кадастрового номера1 9.09.2012
Характеристики объекта
    Адрес (местоположение) г.Санкт-Петербург, Железноводская улица, дом 66, литера А
    Площадь, кв.м8307.1
    Назначение Многоквартирный дом
    Количество этажей 10
    Количество подземных этажей 1
    Материал наружных стен Кирпичные
    Год ввода в эксплуатацию 1968
Сведения о кадастровой стоимости
    Кадастровая стоимость (руб) 227810455.48
    Дата определения 01.01.2018
    Дата внесения 09.01.2019
    
'66:41:0206006:1398': {
    'Общая информация': {
      'Вид объекта недвижимости': 'Помещение',
      'Статус объекта': 'Актуально',
      'Кадастровый номер': '66:41:0206006:1398',
      'Дата присвоения кадастрового номера': '23.01.2014',
      'Форма собственности': 'Частная'
    },
    'Характеристики объекта': {
      'Адрес (местоположение)': 'Свердловская область, г. Екатеринбург, ул. Свердлова, д. 27',
      'Площадь, кв.м': '110.4',
      'Назначение': 'Нежилое',
      'Этаж': 'б/н'
    },
    'Сведения о кадастровой стоимости': {
      'Кадастровая стоимость (руб)': '4322093.83',
      'Дата определения': '01.01.2019',
      'Дата внесения': '10.01.2020'
    },
    'Ранее присвоенные номера': {
      'Условный номер': '66:01/01:00:886:27:21',
      'Инвентарный номер': '0\\4958\\А\\21\\1\\004'
    },
    'Сведения о правах и ограничениях (обременениях)': {
      'Вид, номер и дата государственной регистрации права': 'Собственность\n№ 66-01/01-38/2003-373 от 25.04.2003',
      'Ограничение прав и обременение объекта недвижимости': '№ 66-01/01-264/2003-58 от 20.07.2003\nПрочие ограничения прав и обременения объекта недвижимости\n№ 66-01/01-38/2003-373 от 24.04.2003'
    }
}

{'found': 1,
  'list': [
    {'address': 'Санкт-Петербург, ул. Железноводская, д. 66, литера. А, '
                       'кв. 153',
            'area': '64,3',
            'cad_cost': '7180055,58 руб.',
            'cad_num': '78:06:0220603:3048',
            'cost_definition_date': '01.01.2018',
            'cost_insertion_date': '09.01.2019',
            'floor': '7',
            'notes': [],
            'obj_type': 'Объект капитального строительства',
            'oks_purpose': 'Жилое',
            'oks_type': 'Квартира, Жилое помещение',
            'oks_type_more': 'Помещение',
            'ownersheep_type': '',
            'prev_nums': [{'type': 'cond_num', 'value': '78-78-33/102/2011-25'}],
            'refresh_date': '20.09.2020',
            'reg_date': '19.09.2012',
            'rights': [{'restriction': '',
                        'right': '78-78-33/102/2011-025 от 18.02.2011 '
                                 '(Собственность)'}],
            'rnb': {'bounds': [],
                    'rights': [{'right_description': 'Собственность',
                                'right_origin': '78-78-33/102/2011-025 от '
                                                '18.02.2011 (Собственность)',
                                'right_reg_date': '2011-02-18T00:00:00.000+0000',
                                'right_reg_number': '78-78-33/102/2011-025'}]},
            'status': 'Учтенный',
            'unit': 'кв.м',
            'update_date': '02.10.2023'}],
  'query': 'success'}

"""


def convert_model_to_appropriate_dict(model: SearchResponse) -> dict:
    result = model.results[0]
    return {
    'Общая информация': {
      'Вид объекта недвижимости': result.oks_type_more or result.obj_type,
      'Статус объекта': result.status,
      'Кадастровый номер': result.cad_num,
      'Дата присвоения кадастрового номера': result.reg_date,
      'Форма собственности': result.ownersheep_type,
    },
    'Характеристики объекта': {
      'Адрес (местоположение)': result.address,
      'Площадь, кв.м': result.area,
      'Назначение': result.oks_purpose,
      'Этаж': result.floor,
    },
    'Сведения о кадастровой стоимости': {
      'Кадастровая стоимость (руб)': result.cad_cost,
      'Дата определения': result.cost_definition_date,
      'Дата внесения': result.cost_insertion_date
    },
    'Ранее присвоенные номера': {
      'Условный номер': [el for el in result.prev_nums if el.type == 'cond_num'][0].value if len([el for el in result.prev_nums if el.type == 'cond_num']) > 0 else '',
      'Инвентарный номер': [el for el in result.prev_nums if el.type == 'inv_num'][0].value if len([el for el in result.prev_nums if el.type == 'inv_num']) > 0 else ''
    },
    'Сведения о правах и ограничениях (обременениях)': {
      'Вид, номер и дата государственной регистрации права': '\n'.join([el.right.strip() for el in filter(lambda x: x.right, result.rights)]) if len(result.rights) > 0 else '',
      # 'Ограничение прав и обременение объекта недвижимости': result.rnb.bounds[0].bound_r if len(result.rnb.bounds) > 0 else ''
    }
}


class CadObject:
    cad_id: str
    card: str

    def set_card(self, card: dict):
        res = ''
        for div, subdict in card.items():
            res += div
            for key, string in subdict.items():
                res += key.strip()
                res += string.strip()
        self.card = res.strip()

    def __eq__(self, other: str):
        if not isinstance(other, str):
            return False
        return self.cad_id == other

    def __hash__(self):
        return hash(self.cad_id)


class Room(CadObject):
    pass


class Building(CadObject):
    address: str
    rooms: list[Room]

    def __init__(self):
        super().__init__()
        self.rooms = []

    def find_room(self, cad_id: str):
        return filter(lambda room: room.cad_id == cad_id, self.rooms).__next__()

    def set_room_card(self, room_id: str, card: dict):
        try:
            self.find_room(room_id).set_card(card)
        except ValueError:
            pass


class Land(CadObject):
    building: Building

    def set_id(self, cad_id):
        if cad_id in ('данные отсутствуют', self.building.cad_id):
            return
        self.cad_id = cad_id

    @property
    def exists(self):
        return self.cad_id != 'данные отсутствуют'


class Parser:
    def __init__(self, src):
        self.land = Land()
        self.land.building = Building()
        self.src = src

    def get_objects_from_seventh(self) -> None:
        wb: Workbook = load_workbook(self.src)

        ws_first: Worksheet = wb.get_sheet_by_name("Раздел 1")
        ws_seventh: Worksheet = wb.get_sheet_by_name("Раздел 7")

        self.land.building.cad_id = ws_first["B2"].value
        self.land.building.address = ws_first["B6"].value

        self.land.cad_id = ws_first["B15"].value

        for row in [x for x in ws_seventh.rows][1:]:
            room = Room()
            room.cad_id = row[1].value
            self.land.building.rooms.append(room)

    async def get_objects_from_api(self):
        if self.land.exists:
            land = await api.search_cadastral_full(self.land.cad_id)
            land = convert_model_to_appropriate_dict(land)
            if land:
                self.land.set_card(land)

        building = await api.search_cadastral_full(self.land.building.cad_id)
        building = convert_model_to_appropriate_dict(building)
        if building:
            self.land.building.set_card(building)

        # with ThreadPoolExecutor() as executor:
        #     print(executor._max_workers)
        #     results = list(executor.map(api.search_cadastral_full, (x.cad_id for x in self.land.building.rooms)))
        #     results = [x.model_dump() for x in results]

        futures = [api.search_cadastral_full(x.cad_id) for x in self.land.building.rooms]
        results = await asyncio.gather(*futures)
        pprint.pprint(results[0].model_dump())
        results = [convert_model_to_appropriate_dict(x) for x in results if x]

        for result in results:
            # print(result)
            if result:
                self.land.building.set_room_card(result["Общая информация"]["Кадастровый номер"], result)

    def parse_objects(self):
        if self.land.exists:
            try:
                land = parse_land(self.land.cad_id)
            except TimeoutException:
                print('one more time...')
                land = parse_land(self.land.cad_id)
            if land:
                self.land.set_card(land)

        try:
            building = parse_building(self.land.building.cad_id)
        except TimeoutException:
            print('one more time...')
            building = parse_building(self.land.building.cad_id)
        if building:
            self.land.building.set_card(building)

    def write_objects(self):
        # print(1111, *map(lambda x: x.cad_id, self.land.building.rooms))

        wb: Workbook = load_workbook(self.src)

        ws_first: Worksheet = wb.get_sheet_by_name("Раздел 1")
        ws_seventh: Worksheet = wb.get_sheet_by_name("Раздел 7")

        ws_first["C2"] = self.land.building.card

        if self.land.exists:
            ws_first["C15"] = self.land.card

        # print(*list(map(lambda x: x.cad_id, self.land.building.rooms)))

        for i, row in enumerate([x for x in ws_seventh.rows][1:]):
            cad_id = row[1].value
            # print(cad_id)
            room = self.land.building.find_room(cad_id)
            ws_seventh.cell(row=i + 2, column=8).value = self.land.building.cad_id
            card_cell = ws_seventh.cell(row=i + 2, column=9)
            card_cell.value = room.card
            card_cell.alignment = Alignment(wrapText=True)

        wb.save(self.src)

    async def process_objects(self):
        self.get_objects_from_seventh()
        await self.get_objects_from_api()  # self.parse_objects()
        self.write_objects()


# if __name__ == '__main__':
#     src = input()
#     parser_obj = Parser(src)
#     parser_obj.process_objects()

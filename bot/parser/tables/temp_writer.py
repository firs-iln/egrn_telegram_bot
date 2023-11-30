# TODO: REFACTOR ALL THIS SHIT
import os
import re

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

WB_PATH = '/bot/parser/tables/!_Реестр_МКД_ШАБЛОН 2.xlsx'

statuses = {
    "нежилое": 'НЖ',
    "жилое": "КВ",
    "машиноместо": "ММ",
    "квартира": "КВ",
}


def resize_columns(ws: Worksheet):
    for col in ws.columns:
        set_len = 0
        column = col[0].column_letter
        for cell in col:
            val = cell.value
            if not val:
                continue
            val = val.split('\n')[0]
            new = len(str(val))
            set_len = new if new > set_len else set_len
        set_col_width = set_len + 1
        ws.column_dimensions[column].width = set_col_width


def write_first_and_seventh(addr: str,
                            cad_id: str,
                            first_div: list,
                            seventh_div: list,
                            src: str):
    wb: Workbook = Workbook()
    ws = wb.active
    wb.remove(ws)

    # writing divs
    ws1: Worksheet = wb.create_sheet('Раздел 1')
    ws7: Worksheet = wb.create_sheet('Раздел 7')
    for row in first_div:
        ws1.append([row[0], row[1]])
    for row in seventh_div:
        ws7.append([x for x in row if x])
    resize_columns(ws1)
    resize_columns(ws7)
    print("proceed")
    res = f'bot/files/xlsx/{cad_id}_{addr}.xlsx'
    # os.system(f'touch "{res}"')
    wb.save(res)
    return res


def write_to_table(order_id: int,  # deprecated
                   cad_id: str,
                   addr: str,
                   fias_id: str,
                   first_div: list,
                   seventh_div: list,
                   building: dict[str, dict[str, str]],
                   land: dict[str, dict[str, str]]):
    wb: Workbook = load_workbook(filename=WB_PATH)

    # writing divs
    ws1: Worksheet = wb.create_sheet('Раздел 1')
    ws7: Worksheet = wb.create_sheet('Раздел 7')
    for row in first_div:
        ws1.append([row[0], row[1]])
    for row in seventh_div:
        ws7.append([x for x in row if x])
    resize_columns(ws1)
    resize_columns(ws7)

    # writing building
    ws_building: Worksheet = wb.get_sheet_by_name("Дом")
    ws_building["B2"] = cad_id
    ws_building["B9"] = building['Характеристики объекта'].get('Адрес (местоположение)', addr)
    ws_building["B8"] = fias_id

    cell = ws1.cell(row=2, column=3)
    cell.value = join_room(building)

    cell = ws1.cell(row=15, column=3)
    cell.value = join_room(land)

    res = f'files/xlsx/{cad_id}_{addr}.xlsx'
    wb.save(res)
    return res


def join_room(room: dict) -> str:  # deprecated
    res = ''
    for div, subdict in room.items():
        res += div
        for key, string in subdict.items():
            res += key
            res += string
    return res


def write_rooms_to_seventh(cad_id, src, rooms, row_id):  # deprecated
    wb: Workbook = load_workbook(filename=src)
    ws_seventh: Worksheet = wb.get_sheet_by_name("Раздел 7")

    for i, room in enumerate(rooms.values()):
        if room:
            # print(room)
            res = join_room(room)
            # row = ws_seventh[row_id + 1]
            ws_seventh.cell(row=row_id + 2, column=8).value = cad_id
            ws_seventh.cell(row=row_id + 2, column=9).value = res
    wb.save(src)


top_level_keys = [
    'Общая информация',
    'Характеристики объекта',
    'Сведения о кадастровой стоимости',
    'Сведения о правах и ограничениях (обременениях)',
]


def unpack_card(text: str) -> dict:  # deprecated
    text = [x for x in re.split('|'.join(top_level_keys), text) if x]
    res = {}
    for div in text:
        cur_items = []
        cur_item = ''
        for letter in div:
            if letter == letter.upper() or not letter.isalpha():
                if len(cur_items) == 2:
                    res[cur_item[0]] = cur_item[1]
                    cur_item = letter
                    cur_items = []
                else:
                    cur_items.append(cur_item)
                    cur_item = letter

    print(*text, sep='\n')


def find_property(text: str, sfrom: str, sto: str) -> str:  # deprecated
    return text[text.find(sfrom) + len(sfrom):text.find(sto)]


def from_seventh_to_rooms_one(src, row_id):  # deprecated
    wb: Workbook = load_workbook(src)
    row_id += 1
    ws_rooms: Worksheet = wb.get_sheet_by_name("Помещения")
    ws_seventh: Worksheet = wb.get_sheet_by_name("Раздел 7")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    card_cell = ws_seventh.cell(row=row_id, column=9)
    room_info: str = card_cell.value
    if not room_info:
        return None

    cad_id = find_property(room_info, "Кадастровый номер", "Дата присвоения кадастрового номера")
    cad_cell = ws_rooms.cell(row=row_id, column=5)
    cad_cell.value = cad_id

    new_card_cell = ws_rooms.cell(row=row_id, column=7)
    new_card_cell.value = ''.join(room_info.split())
    new_card_cell.alignment = Alignment(wrapText=True)

    purpose = find_property(room_info, 'Назначение', 'Этаж')
    status = statuses.get(purpose.lower(), '')
    status_cell = ws_rooms.cell(row=row_id, column=1)
    status_cell.value = status

    address = find_property(room_info, 'Адрес (местоположение)', 'Площадь, кв.м')
    address_cell = ws_rooms.cell(row=row_id, column=2)
    address_cell.value = address

    number = re.split('кв\. |пом\. ', address)[-1]
    number_cell = ws_rooms.cell(row=row_id, column=3)
    number_cell.value = number

    area_to_return = find_property(room_info, 'Площадь, кв.м', 'Назначение')
    area = area_to_return.replace('.', ',')
    formal_area_cell = ws_rooms.cell(row=row_id, column=4)
    formal_area_cell.value = area
    real_area_cell = ws_rooms.cell(row=row_id, column=8)
    real_area_cell.value = area

    floor = find_property(room_info, 'Этаж', 'Сведения о кадастровой стоимости')
    floor_cell = ws_rooms.cell(row=row_id, column=12)
    floor_cell.value = floor if floor else 'б/н'

    for cell in ws_rooms[f"{row_id}:{row_id}"]:
        cell.border = thin_border

    if row_id == 2:
        ws_rooms.freeze_panes = 'B1'

    wb.save(src)

    return float(area_to_return) if status in ('КВ', 'НЖ', 'ММ') else None


def write_areas_parts(src, total_area: float, areas_by_row: dict[int, float]) -> None:  # deprecated
    wb: Workbook = load_workbook(src)
    ws_rooms: Worksheet = wb["Помещения"]
    parts_by_row = {key + 1: value / total_area * 100 for key, value in areas_by_row.items()}
    for row_id, part in parts_by_row.items():
        part_cell = ws_rooms.cell(row=row_id, column=6)
        part_cell.value = part
    wb.save(src)


def write_rooms(src, rooms):  # deprecated
    # print(rooms)
    wb: Workbook = load_workbook(filename=src)
    ws_rooms: Worksheet = wb.get_sheet_by_name("Помещения")
    for i, room in enumerate(rooms.values()):
        if room:
            ws_rooms.cell(row=i + 1, column=1).value = statuses.get(
                room['Характеристики объекта'].get('Назначение', ''), '')
            ws_rooms.cell(row=i + 1, column=2).value = room['Характеристики объекта'].get('Адрес (местоположение)', '')
    wb.save(src)


if __name__ == '__main__':
    wb: Workbook = load_workbook(
        '/bot/files/xlsx/ 664102060302313_Свердловская область, г. Екатеринбург, ул. Свердлова, д. 27.xlsx')
    ws: Worksheet = wb["Помещения"]
    unpack_card(ws["G3"].value)

import os

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import re
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from collections import defaultdict
from copy import copy

import dotenv

dotenv.load_dotenv()

from bot.parser.parse_website.dadata_rep import dadata

statuses = {
    "нежилое": 'НЖ',
    "жилое": "КВ",
    "машиноместо": "ММ",
    "квартира": "КВ"
}

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

WB_PATH = 'bot/parser/tables/!_Реестр_МКД_ШАБЛОН 2.xlsx'


def get_addr_and_cad_id(src):
    wb: Workbook = load_workbook(src)
    ws_first: Worksheet = wb.get_sheet_by_name("Раздел 1")

    return ws_first["B2"].value, ws_first["B6"].value


def check_for_spaces_in_column(column_letter: str, src: str) -> bool:
    wb: Workbook = load_workbook(src)
    ws: Worksheet = wb.get_sheet_by_name("Раздел 7")
    for cell in ws[column_letter]:
        if cell.value is None:
            continue
        if ' ' in cell.value:
            return True
    return False


def get_rooms_content_from_xlsx_with_cards(src):
    rooms = {}

    wb: Workbook = load_workbook(src)

    ws_first: Worksheet = wb.get_sheet_by_name("Раздел 1")
    rooms[ws_first["B2"].value] = ws_first["C2"].value  # get building card
    rooms[ws_first["B15"].value] = ws_first["C15"].value  # get land card

    ws_seventh: Worksheet = wb.get_sheet_by_name("Раздел 7")
    for row in list(ws_seventh.rows)[1:]:
        rooms[row[1].value] = row[8].value

    # print(rooms.keys())

    return {key: value for key, value in rooms.items() if key != 'данные отсутствуют' and value is not None}


def find_property(text: str, sfrom: str, sto: str, mod: int = 0) -> str:
    if sto in text:
        return text[text.find(sfrom) + len(sfrom):text.find(sto) + mod]
    else:
        return text[text.find(sfrom) + len(sfrom):]


def make_register_file(src, order_id: int = 10002):
    cad, address = get_addr_and_cad_id(src)

    rooms = get_rooms_content_from_xlsx_with_cards(src)

    wb: Workbook = load_workbook(WB_PATH)

    ws_building: Worksheet = wb.get_sheet_by_name("Дом")
    ws_building.cell(row=1, column=2).value = order_id
    ws_building.cell(row=2, column=2).value = cad
    ws_building.cell(row=9, column=2).value = address
    ws_building.cell(row=8, column=2).value = dadata.get_fias_id(address)

    ws_rooms: Worksheet = wb.get_sheet_by_name("Помещения")
    ws_registry: Worksheet = wb.get_sheet_by_name("Реестр")

    # noinspection PyProtectedMember
    base_style = ws_registry["C4"]._style

    total_area = 0
    areas_by_row = {}
    row_id = 2

    res = f'bot/files/xlsx/{order_id}_Реестр_{address}.xlsx'

    # print(*rooms.items(), sep="\n")

    for cad_id, room_info in rooms.items():
        # print(cad_id, room_info)
        address = find_property(room_info, 'Адрес (местоположение)', 'Площадь, кв.м')

        kind = find_property(room_info, 'Вид объекта недвижимости', 'Статус объекта').lower()
        if kind == 'помещение':
            purpose = find_property(room_info, 'Назначение', 'Этаж')
            if 'Вид, номер и дата государственной регистрации права' not in room_info:
                status = 'ОИ'
            if "лк" in address.lower():
                status = "ЛК"
            else:
                status = statuses.get(purpose.lower(), '')
            number = re.split(r'кв\.|пом\.|д\.|дом|квартира|кв\. ', address)[-1]
            number_cell = ws_rooms.cell(row=row_id, column=3)
            number_cell.value = number
            if 'б/н' in number:
                status = 'ОИ'
            elif 'Н' in number:
                status = 'НЖ'
        elif kind == 'здание':
            status = "МКД"
        elif kind == 'земельный участок':
            status = 'ЗУ'
        status_cell = ws_rooms.cell(row=row_id, column=1)
        status_cell.value = status

        cad_id = find_property(room_info, "Кадастровый номер", "Дата присвоения кадастрового номера")
        cad_cell = ws_rooms.cell(row=row_id, column=5)
        cad_cell.value = cad_id

        new_card_cell = ws_rooms.cell(row=row_id, column=7)
        new_card_cell.value = room_info
        # new_card_cell.alignment = Alignment(wrapText=True)

        address_cell = ws_rooms.cell(row=row_id, column=2)
        address_cell.value = address

        if status != 'ЗУ':
            area_to_return = find_property(room_info, 'Площадь, кв.м', 'Назначение')
        else:
            area_to_return = find_property(room_info, 'Площадь, кв.м', 'Категория земель')

        area = area_to_return.replace('.', ',')
        formal_area_cell = ws_rooms.cell(row=row_id, column=4)
        formal_area_cell.value = area
        real_area_cell = ws_rooms.cell(row=row_id, column=8)
        real_area_cell.value = area

        if status in statuses.values():
            floor = find_property(room_info, 'Этаж', 'Сведения о кадастровой стоимости')
            floor_cell = ws_rooms.cell(row=row_id, column=12)
            floor_cell.value = floor if floor else 'б/н'

        # for cell in ws_rooms[f"{row_id}:{row_id}"]:
        #     cell.border = thin_border

        if row_id == 2:
            ws_rooms.freeze_panes = 'A2'

        area = float(area_to_return.replace(',', '.')) if status in ('КВ', 'НЖ', 'ММ') else None
        if area:
            total_area += area
            areas_by_row[row_id] = area

        if not status:
            status = find_property(room_info, 'Вид объекта недвижимости', 'Статус объекта')

        prop = find_property(room_info,
                             'Вид, номер и дата государственной регистрации права',
                             'Ограничение прав и обременение объекта недвижимости',
                             )

        # print(prop if status == 'НЖ' else None)
        # print(cad_id)
        # print(prop)
        if status in ('КВ', 'НЖ'):
            if 'долевая' in prop:
                prop = prop.split('Общая долевая собственность')
                regs = list(map(lambda x: ["Общая долевая собственность", *x.split("от")], prop[1:]))
                ws_rooms.cell(row=row_id, column=9).value = f'П+{len(regs)}'
            elif 'Собственность' in prop:
                ws_rooms.cell(row=row_id, column=9).value = f'П+{1}'

        wb.save(res)

        row_id += 1

    parts_by_row = {key: round(value / total_area * 100, 3) for key, value in areas_by_row.items()}
    for row_id, part in parts_by_row.items():
        status = ws_rooms.cell(row=row_id, column=1).value
        if status in ('КВ', 'НЖ', 'ММ'):
            ws_rooms.cell(row=row_id, column=6).value = part

            # ws_registry.cell(row=row_id - 2, column=18).value = part

    base_style = ws_rooms["B2"]._style

    wb.save(res)

    # print(ws_rooms["C2"].value)
    # print(ws_rooms.cell(row=2, column=3).value)
    # print(ws_rooms["C3"].value)
    # print(ws_rooms["C4"].value)

    sort_rows(ws_rooms, start=3)

    wb.save(res)

    make_registry_sheet(wb, res, ws_rooms, ws_registry)

    for row in ws_rooms.iter_rows(min_row=2):
        for cell in row:
            cell._style = base_style

    for row in ws_registry.iter_rows(min_row=2):
        for cell in row:
            cell._style = base_style

    for i, row in enumerate(ws_rooms.iter_rows(min_row=2), start=2):
        ws_rooms.row_dimensions[i].height = 12

    wb.save(res)

    return res, total_area


def sort_rows(ws: Worksheet, start: int = 1) -> None:
    rows = list(ws.iter_rows(min_row=start + 1))

    # print(rows[])

    # print(ws.cell(row=start, column=3).value)
    # print(ws.cell(row=start + 1, column=3).value)

    # print(ws["C2"].value)
    # print(ws["C3"].value)
    # print(ws["C4"].value)
    # print(*map(lambda x: x.value, rows[0]))
    ws.delete_rows(start + 1, len(rows))
    rows.sort(key=lambda x: (x[0].value, float(x[2].value) if x[2].value.strip().isdigit() else float('inf')))
    for i, row in enumerate(rows, start=start + 1):
        for cell in row:
            ws.cell(row=i, column=cell.col_idx).value = cell.value


def make_registry_sheet(wb: Workbook, src: str, ws_rooms: Worksheet, ws_registry: Worksheet):
    row_id = 2
    for row in ws_rooms.iter_rows(min_row=4):
        status = row[0].value
        is_room = status not in ('МКД', 'ЗУ', 'ЛК')
        if is_room:
            room_info = row[6].value
            # print(room_info)
            cad_id = row[4].value
            number = row[2].value
            area = row[3].value
            part = row[5].value
            owner_type_cell = ws_registry.cell(row=row_id, column=10)

            owner_type_cell.value = 'ФЛ' if status in ('КВ', 'ММ') else 'ОИ'
            # print(status, owner_type_cell.value)
            # print(f'room {row_id} {status} {cad_id} {number} {area}')
            ws_registry.cell(row=row_id, column=1).value = status
            ws_registry.cell(row=row_id, column=2).value = cad_id
            ws_registry.cell(row=row_id, column=3).value = number
            ws_registry.cell(row=row_id, column=4).value = area
            if status not in ('НЖ', 'ОИ'):
                ws_registry.cell(row=row_id, column=17).value = area

            prop = find_property(room_info,
                                 'Вид, номер и дата государственной регистрации права',
                                 'Ограничение прав и обременение объекта недвижимости',
                                 # mod=1
                                 )
            if 'долевая' in prop:
                prop = prop.split('\n')
                prop = list(map(lambda x: x.replace('(Общая долевая собственность)', '').strip(), prop))
                regs = list(map(lambda x: ["Общая долевая собственность", *x.split("от")], prop))
                value = ', '.join([' от '.join((x[1].strip(), x[2].strip())) for x in regs])
                ws_registry.cell(row=row_id, column=13).value = regs[0][0]
                ws_registry.cell(row=row_id, column=14).value = f'Долевая собственность {value}'
            elif 'Собственность' in prop:
                regs = ['Собственность', *prop.replace('(Cобственность)', '').strip().split('от')]
                # print(regs)
                ws_registry.cell(row=row_id, column=13).value = regs[0]
                ws_registry.cell(row=row_id, column=14).value = regs[1].replace('№', '').strip()
                ws_registry.cell(row=row_id, column=15).value = regs[2].strip()

            no_data_placeholder = 'нет данных'

            ws_registry.cell(row=row_id, column=7).value = no_data_placeholder
            ws_registry.cell(row=row_id, column=8).value = no_data_placeholder
            ws_registry.cell(row=row_id, column=9).value = no_data_placeholder

            ws_registry.cell(row=row_id, column=5).value = 1
            ws_registry.cell(row=row_id, column=6).value = 1
            if status not in ('НЖ', 'ОИ'):
                ws_registry.cell(row=row_id, column=18).value = part

            row_id += 1

    # wb.save(f'src.xlsx')
    wb.save(src)


def parse_owners_by_cad_id(src: str, column_map: dict):
    wb: Workbook = load_workbook(src)
    print(src, wb.sheetnames)
    ws = wb['Реестр']

    A = ord('A')

    cad_id_column = ord(column_map['cad_id']) - A
    owner_columns = list(map(lambda x: ord(x) - A, column_map['owner'].split(',')))
    extract_column = ord(column_map['extract']) - A if 'extract' in column_map else None
    registred_columns = list(map(lambda x: ord(x) - A, column_map['registred'].split(',')))
    # print(list(registred_columns), column_map['registred'])
    part_columns = list(map(lambda x: ord(x) - A, column_map['part'].split(',')))
    print(part_columns)


    owners_by_cad_id = defaultdict(list)

    for row in ws.iter_rows(min_row=7):
        cad_id = row[cad_id_column].value
        # print(11111, list(owner_columns))
        # print(*map(lambda x: x.value, row))

        owners = ''
        for owner_column in owner_columns:
            owner = ' '.join([owners, row[owner_column].value])

        if extract_column:
            extract = row[extract_column].value

        registred = ''
        for registred_column in registred_columns:
            registred = ' '.join([registred, row[registred_column].value])

        has_date = 'от' in registred
        if not has_date:
            continue

        if len(part_columns) == 1:
            value = row[part_columns[0]].value
            print(value)
            part = value if value != '1' else '1/1'
        else:
            value = ''
            for part_column in part_columns:
                value = '/'.join([value, row[part_column].value])
            part = value if value != '/1' else '1/1'

        if extract_column:
            extract = row[extract_column].value
            owners_by_cad_id[cad_id].append([owner, registred, part, extract])
        else:
            owners_by_cad_id[cad_id].append([owner, registred, part])

    print(*owners_by_cad_id.items(), sep='\n')
    return owners_by_cad_id


def join_owners_and_registry(registry_src: str,
                             owners_src: str,
                             column_map: dict):
    owners_by_cad_id = parse_owners_by_cad_id(owners_src, column_map)

    wb: Workbook = load_workbook(registry_src)
    ws: Worksheet = wb['Реестр']

    rows_to_insert = []
    rows_to_delete = []

    # noinspection PyProtectedMember
    base_style = ws["C4"]._style

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cad_id = row[1].value.strip()
        status = row[0].value
        print(1111, cad_id, cad_id in owners_by_cad_id)
        owners = owners_by_cad_id[cad_id]
        print(2222, owners)
        if status in ('НЖ', 'ОИ'):
            continue
        if len(owners) == 1:
            print()
            print('one owner', owners)
            owner, registred, part = owners[0]
            owner = owner.strip()
            registred = registred.strip()
            registred = re.split(' от |№', registred)

            table_date = ws.cell(row=i, column=15).value.strip()
            # has_date = re.match('\d{2}\.\d{2}\.\d{4}', table_date)
            # if not has_date:
            #     continue
            print("yoy?", registred)
            owner = owner.split(' ')
            print(registred, table_date)
            if table_date == registred[-1]:
                print(i, owner, 'yes!!!!!!!!!!!!')
                ws.cell(row=i, column=7).value = owner[0]
                ws.cell(row=i, column=8).value = owner[1]
                ws.cell(row=i, column=9).value = ' '.join(owner[2:])
            else:
                new_row = [*[x.value for x in row]]
                new_row[6] = owner[0]
                new_row[7] = owner[1]
                new_row[8] = ' '.join(owner[2:])
                new_row[12] = registred[0]
                new_row[13] = registred[1]
                new_row[14] = registred[2]
                rows_to_insert.append(new_row)
        else:
            for owner, registred, part in owners:
                print("yoy", registred, owner)
                owner = owner.strip()
                registred = registred.strip()
                registred = re.split(' от | №', registred)
                print(registred)
                owner = owner.split()
                new_row = [*[x.value for x in row]]
                part = part.split('/')
                new_row[4] = part[0]
                new_row[5] = part[1]
                ratio = int(part[0]) / int(part[1])
                print("ratio", ratio)
                new_row[16] = str(round(float(row[16].value.replace(",", ".")) * ratio, 3)).replace('.', ',')
                print(new_row[16])
                new_row[17] = str(round(float(row[17].value) * ratio, 3)).replace('.', ',')
                print(new_row[17])
                new_row[6] = owner[0]
                new_row[7] = owner[1]
                new_row[8] = ' '.join(owner[2:])
                new_row[12] = registred[0]
                new_row[13] = registred[1].strip()
                new_row[14] = registred[2].strip()
                rows_to_insert.append(new_row)

    # wb.save(registry_src)

    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i == 9:
            print(i, row[6].value.strip())
        if row[6].value.strip() == 'нет данных':
            print(i, row[6].value)
            print(*map(lambda x: x.value, row))
            rows_to_delete.append(i - 1)

    new_rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[6].value.strip() != 'нет данных' or row[0].value.strip() != 'КВ':
            print('save', i)
            new_rows.append([x.value for x in row])
        else:
            print(i, row[6].value.strip())

    ws.delete_rows(2, amount=ws.max_row - 1)

    for row in new_rows:
        print(row)
        ws.append(row)

    for row in rows_to_insert:
        ws.append(row)

    sort_rows(ws)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell._style = base_style

    # parts_by_cad_id = {}
    # for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    #     if row[1].value not in parts_by_cad_id:
    #         parts_by_cad_id[row[1].value] = [(int(row[4].value), int(row[5].value))]
    #         continue
    #     parts_by_cad_id[row[1].value].append((int(row[4].value), int(row[5].value)))
    #
    # invalid_cads = []
    # for cad, parts in parts_by_cad_id.items():
    #     print(parts)
    #     if sum([x[0] for x in parts]) != max([x[1] for x in parts]):
    #         invalid_cads.append(cad)
    #
    # print(invalid_cads)
    # r, w = Font(), Font(color="FF0000", bold=True)
    # red = InlineFont(color='FF000000', b=True)
    # for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    #     if row[1].value in invalid_cads:
    #         # if cell.column in (5, 6):
    #         rich_string1 = CellRichText([TextBlock(red, row[4].value), ])
    #         rich_string2 = CellRichText(TextBlock(red, row[5].value), )
    #
    #         print(row[1].value, ws[f'E{i}'])
    #         ws[f'E{i}'] = rich_string1
    #         ws[f'F{i}'] = rich_string2

    ws.column_dimensions['O'].width += 3

    wb.save(registry_src)
    return registry_src


if __name__ == '__main__':
    res = make_register_file(
        "/bot/files/xlsx/780602206032746_г_Санкт_Петербург,_Железноводская_улица,_дом_66,.xlsx")
    join_owners_and_registry("10002_Реестр_г.Санкт-Петербург, Железноводская улица, дом 66, литера А.xlsx",
                             "Готовый_Росреестр_Санкт_Петербург,_ул_Железноводская,_д_66,_литера (2).xlsx",
                             {'cad_id': 'B', 'owner': 'D', 'registred': 'I', 'part': 'F'})
